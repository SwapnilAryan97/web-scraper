from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from .config import Settings
from .models import ImageJob
from .utils import normalize_header


def _resolve_required_column(
    available_headers: dict[str, int],
    preferred_name: str,
    aliases: list[str],
    *,
    label: str,
) -> str:
    candidates = [preferred_name, *aliases]
    for candidate in candidates:
        normalized = normalize_header(candidate)
        if normalized in available_headers:
            return normalized
    raise ValueError(f"Could not resolve required {label} column from {candidates}")


def load_image_jobs(sheet_path: Path, settings: Settings) -> list[ImageJob]:
    workbook = load_workbook(sheet_path, data_only=True)
    worksheet = (
        workbook[settings.sheet.sheet_name]
        if settings.sheet.sheet_name
        else workbook.active
    )

    header_row = settings.sheet.header_row
    headers_by_index: dict[int, str] = {}
    normalized_headers: dict[str, int] = {}

    for row in worksheet.iter_rows(min_row=header_row, max_row=header_row):
        for cell in row:
            if cell.value is None:
                continue
            raw_header = str(cell.value).strip()
            normalized = normalize_header(raw_header)
            if not normalized:
                continue
            headers_by_index[cell.column] = raw_header
            normalized_headers[normalized] = cell.column

    if not headers_by_index:
        raise ValueError("No headers found in the configured sheet")

    sku_key = _resolve_required_column(
        normalized_headers,
        settings.sheet.sku_column,
        settings.sheet.sku_aliases,
        label="SKU",
    )
    product_name_key = _resolve_required_column(
        normalized_headers,
        settings.sheet.product_name_column,
        settings.sheet.product_name_aliases,
        label="product name",
    )

    image_slot_patterns = [
        re.compile(pattern, flags=re.IGNORECASE)
        for pattern in settings.sheet.image_slot_patterns
    ]
    image_headers = {
        normalize_header(raw_header): raw_header
        for raw_header in headers_by_index.values()
        if any(
            pattern.match(normalize_header(raw_header))
            for pattern in image_slot_patterns
        )
    }

    jobs: list[ImageJob] = []
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        row_values: dict[str, object] = {}
        row_is_empty = True

        for column_index, value in enumerate(row, start=1):
            raw_header = headers_by_index.get(column_index)
            if raw_header is None:
                continue
            normalized_header = normalize_header(raw_header)
            row_values[normalized_header] = value
            if value not in (None, ""):
                row_is_empty = False

        if row_is_empty:
            continue

        sku_raw = row_values.get(sku_key)
        product_name_raw = row_values.get(product_name_key)
        sku = str(sku_raw).strip() if sku_raw not in (None, "") else ""
        product_name = (
            str(product_name_raw).strip() if product_name_raw not in (None, "") else ""
        )
        if not sku or not product_name:
            continue

        metadata = {
            key: str(value).strip()
            for key, value in row_values.items()
            if value not in (None, "")
        }

        for normalized_attr, raw_attr in image_headers.items():
            current_value = row_values.get(normalized_attr)
            if not settings.sheet.process_filled_slots and current_value not in (
                None,
                "",
            ):
                continue
            jobs.append(
                ImageJob(
                    row_number=row_number,
                    sku=sku,
                    product_name=product_name,
                    attribute_name=raw_attr,
                    current_value=(
                        str(current_value).strip()
                        if current_value not in (None, "")
                        else None
                    ),
                    metadata=dict(metadata),
                )
            )

    return jobs
